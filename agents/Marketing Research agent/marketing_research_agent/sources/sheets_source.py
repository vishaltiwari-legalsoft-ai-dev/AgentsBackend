"""Google Sheets data source — pulls Legal Soft's live performance tracker.

The tracker is a transposed monthly grid: metric names down column A, month
columns in ``(Performance)`` / ``(Investment)`` pairs, with quarter and YTD
rollups interleaved. One tab covers a brand and may contain several channel
blocks (e.g. a META block plus a ``GOOGLE`` sub-block).

Auth uses Application Default Credentials with the read-only Drive scope (already
enabled for this project), pulling each tab through the authenticated CSV-export
endpoint — so the Google Sheets API does not need to be enabled. The fetcher is
injectable so ``parse_tracker`` and this class can be unit-tested offline.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import threading
from datetime import date
from typing import Callable, Sequence

from .. import config
from ..schemas import CampaignMetric, DataGap, DateRange, Lead
from .base import CredentialMissingError

logger = logging.getLogger("agentos.mr.sheets")

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Channel keywords that mark a block header row (exact, case-insensitive label).
# Also how a tab title is classified. A vendor whose channel is missing here is
# not a cosmetic problem: see `default_channel` in _find_blocks.
_CHANNEL_HEADERS = {
    "google": "Google", "meta": "META", "facebook": "META", "email": "Email",
    "website": "Websites", "websites": "Websites", "organic": "Organic",
    "linkedin": "LinkedIn",
    # 2026-08 vendor tabs — without these their tabs classified as "Total".
    "chatgpt": "ChatGPT", "microsoft": "Microsoft", "bing": "Microsoft",
    "twitter": "Twitter", "youtube": "YouTube", "tiktok": "TikTok",
}

# Acceptable exact row labels (stripped, case-insensitive) per canonical field.
# Order = priority; first match within a block wins.
_FIELD_LABELS = {
    "spend": ["spend"],
    "leads": ["leads"],
    "qualified_leads": ["qualified leads"],
    "demos_booked": [
        "total demos booked (sdr+vapi+direct)",
        "total demos booked",
        "qualified demos booked (sdr+vapi+direct)",
        "qualified demos booked",
    ],
    "demos_completed": [
        "demos completed (sdr+vapi+direct)",
        "total demos completed (direct)",
        "total demos completed",
    ],
}


def _num(s: str) -> float | None:
    s = (s or "").strip()
    if s in ("", "#N/A", "#DIV/0!", "#REF!", "-", "#VALUE!", "#NAME?"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("%", "").replace("(", "").replace(")", "")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def scan_month_columns(header: list[str]) -> tuple[list[tuple[int, int, int]], list[int]]:
    """``(columns, repeated_months)`` for a tracker header row.

    ``columns`` is (month_num, perf_col, inv_col) per real month; quarter and YTD
    rollup columns are skipped and ``inv_col`` is -1 when a month has no
    Investment column.

    **The FIRST occurrence of a month wins, and the repeat is reported.** This
    used to keep the last one, which is how a second month-column band added to
    the right of the grid silently re-pointed every figure: on 2026-07-27 the
    Overall tab read July spend $48,005.09, and after a layout change the same
    row on the same tab read $5,579.58 — a roll-up *below* the sum of the vendor
    tabs it aggregates. Worse, Performance and Investment were resolved
    independently, so a pair could straddle two different tables. The primary
    grid is the leftmost one; anything repeated to its right is a secondary view
    the parser has no basis for preferring, so it is ignored *and named* —
    ``repeated_months`` is what turns this from a silent swap into a data gap.
    """
    perf: dict[int, int] = {}
    inv: dict[int, int] = {}
    repeated: set[int] = set()
    for col, raw in enumerate(header):
        text = (raw or "").strip().lower()
        if "ytd" in text or "quarter" in text:
            continue
        m = re.match(r"([a-z]+)", text)
        if not m or m.group(1) not in _MONTHS:
            continue
        month = _MONTHS[m.group(1)]
        if "performance" in text:
            target = perf
        elif "investment" in text:
            target = inv
        else:
            continue
        if month in target:
            repeated.add(month)
            continue  # first band wins
        target[month] = col
    out = []
    for month in sorted(perf):
        i = inv.get(month, -1)
        # An Investment column left of its Performance column cannot be its
        # pair — it belongs to another band. Better no billed figure than one
        # lifted from a different table.
        if i < perf[month]:
            i = -1
        out.append((month, perf[month], i))
    return out, sorted(repeated)


def _month_columns(header: list[str]) -> list[tuple[int, int, int]]:
    """Columns only — see :func:`scan_month_columns` for the repeat report."""
    return scan_month_columns(header)[0]


_MONTH_NAMES = ("", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December")


def _repeat_gap_message(title: str, repeated: list[int]) -> str:
    names = ", ".join(_MONTH_NAMES[m] for m in repeated)
    return (f"'{title}': {names} appear(s) in more than one column band — read "
            f"the leftmost grid and ignored the repeat(s). Check the tab layout.")


def _block_channel(title_or_label: str) -> str | None:
    text = (title_or_label or "").strip().lower()
    for key, channel in _CHANNEL_HEADERS.items():
        if re.search(rf"\b{re.escape(key)}\b", text):
            return channel
    return None


def _brand_from_title(title: str) -> str:
    """`Meta 360 RA` -> `360 RA`; strip a leading channel word if present."""
    parts = (title or "").strip().split(None, 1)
    if parts and _block_channel(parts[0]):
        return parts[1].strip() if len(parts) > 1 else title.strip()
    return (title or "").strip()


def _find_blocks(rows: list[list[str]], title: str,
                 default_channel: str = "Total") -> list[tuple[str, int, int]]:
    """Return (channel, start_row, end_row_exclusive) for each channel block.

    The top block's channel comes from the tab title; sub-blocks are introduced
    by a bare channel-name row (e.g. ``GOOGLE``).

    ``default_channel`` is what an unrecognised title falls back to. "Total" is
    right for the consolidated roll-up (A1 = "All"/"Overall") and WRONG for a
    vendor tab — a vendor whose channel is not in ``_CHANNEL_HEADERS``
    ("DanteAgency ChatGPT") landed in a channel called "Total", which the report
    builder pops out and treats as the whole report's totals. Two small vendor
    tabs then WERE the headline KPI strip. Callers reading a vendor tab pass
    "Other" so an unknown channel stays a channel."""
    headers: list[tuple[str, int]] = []
    top_channel = _block_channel(title) or default_channel
    headers.append((top_channel, 1))  # data starts just under the title row
    for i, row in enumerate(rows):
        if i == 0:
            continue
        label = (row[0] if row else "").strip()
        if label and label.lower() in _CHANNEL_HEADERS and len(label.split()) == 1:
            headers.append((_CHANNEL_HEADERS[label.lower()], i + 1))
    headers.sort(key=lambda h: h[1])
    blocks = []
    for idx, (channel, start) in enumerate(headers):
        end = headers[idx + 1][1] - 1 if idx + 1 < len(headers) else len(rows)
        blocks.append((channel, start, end))
    return blocks


def _row_for(rows: list[list[str]], start: int, end: int, candidates: Sequence[str],
             occurrence: int = 1) -> int | None:
    """Index of the row this field should be read from, or None.

    ``candidates`` keeps its meaning: label priority, exact stripped-lowercase
    match, first one that resolves wins.

    ``occurrence`` (1-based) says WHICH row wins when the same label repeats
    inside the block — the convention ``_TEAM_MAP``/``_canonical_block`` in
    snapshots.py already uses, carried in the same tuple position. A candidate
    that appears fewer than ``occurrence`` times is simply not a match, so the
    next candidate is tried and an unmatched field stays absent (never 0).

    The default is occurrence 1, which is byte-for-byte the previous
    first-match-wins behaviour — every existing caller resolves the row it
    always resolved.
    """
    for want in candidates:
        hits = 0
        for i in range(start, end):
            label = (rows[i][0] if i < len(rows) and rows[i] else "").strip().lower()
            if label == want:
                hits += 1
                if hits == occurrence:
                    return i
    return None


def parse_tracker(rows: list[list[str]], year: int, brand: str | None = None) -> tuple[list[CampaignMetric], list[DataGap]]:
    """Parse a tracker tab into monthly ``CampaignMetric`` rows (one per channel
    block per month with data)."""
    if not rows:
        return [], [DataGap("sheets", "empty tab")]
    title = (rows[0][0] if rows[0] else "").strip()
    brand = brand or _brand_from_title(title)
    months, repeated = scan_month_columns(rows[0])
    if not months:
        return [], [DataGap("sheets", f"no month columns found in '{title}'")]

    metrics: list[CampaignMetric] = []
    gaps: list[DataGap] = []
    if repeated:
        gaps.append(DataGap("sheets", _repeat_gap_message(title, repeated)))
    # A vendor tab's own scope. Anything else in the tab is a sub-block.
    own_channel = _block_channel(title) or ("Total" if is_rollup_scope(rows) else "Other")
    for channel, start, end in _find_blocks(rows, title, default_channel=own_channel):
        # A channel that has its own dedicated tab is pasted into vendor tabs as
        # a reference block. Ingesting it from there is wrong twice over: the
        # rows are attributed to the wrong vendor, and the same block is counted
        # once per tab carrying it. On 2026-08-15 the live workbook had the
        # Websites block in SEVEN vendor tabs plus its own — $8,632/month
        # counted eight times ($60,424 of phantom spend a month) — and the two
        # vendors whose own block was still empty showed nothing BUT Websites'
        # numbers under their own name.
        if channel in config.NON_MEDIA_CHANNELS and channel != own_channel:
            gaps.append(DataGap("sheets", (
                f"{brand}: ignored a '{channel}' block pasted into this tab — "
                f"those rows belong to the '{channel}' tab and would be counted twice")))
            continue
        idx = {f: _row_for(rows, start, end, labels) for f, labels in _FIELD_LABELS.items()}
        if idx["spend"] is None and idx["leads"] is None:
            continue  # not a data block
        for missing in [f for f, i in idx.items() if i is None]:
            gaps.append(DataGap("sheets", f"{brand}/{channel}: no '{missing}' row"))

        def val(field: str, perf: int, inv: int) -> float | None:
            i = idx[field]
            if i is None:
                return None
            row = rows[i]
            cell = lambda c: row[c] if 0 <= c < len(row) else ""
            # Official basis (decision 2026-07-27): the team reads the sheet's
            # Performance columns, so every figure — spend included — is
            # Performance first, Investment (billed) fallback.
            return _num(cell(perf)) if _num(cell(perf)) is not None else _num(cell(inv))

        for month, perf, inv in months:
            spend = val("spend", perf, inv)
            leads = val("leads", perf, inv)
            if not spend and not leads:
                continue
            metrics.append(CampaignMetric(
                channel=channel,
                campaign=f"{brand} · {channel}",
                utm_source=channel.lower(),
                utm_medium="paid",
                utm_campaign=brand,
                spend=spend or 0.0,
                leads=int(leads or 0),
                qualified_leads=int(val("qualified_leads", perf, inv) or 0),
                demos_booked=int(val("demos_booked", perf, inv) or 0),
                demos_completed=int(val("demos_completed", perf, inv) or 0),
                date=date(year, month, 1),
            ))
    return metrics, gaps


# --- Authenticated fetch ---------------------------------------------------

DRIVE_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
SHEETS_SCOPE = "https://www.googleapis.com/auth/spreadsheets.readonly"

# --- Deadlines -------------------------------------------------------------
# Every Google call here runs inside a *sync* FastAPI handler, which means it
# runs on anyio's worker threadpool — 40 slots for the whole process. A call
# that stalls does not merely fail slowly, it takes a slot out of circulation;
# ~40 stalled sheet pulls wedge the service including /api/health. So each
# transport gets a deadline we chose, not one we inherited.
#
# googleapiclient does apply an implicit 60s socket timeout of its own
# (`googleapiclient.http.DEFAULT_HTTP_TIMEOUT_SEC`), so these calls were not
# literally unbounded — but 60s is far longer than any healthy Sheets call, it
# is invisible at the call site, and it is a *socket* timeout (per read), so a
# server dribbling bytes can still outlive it. Stating the number here makes it
# reviewable and tunable.
SHEETS_TIMEOUT_SECONDS = 30
# Token mint/refresh against oauth2.googleapis.com (or the metadata server): a
# tiny request with no reason to be slow. google-auth's own default is 120s.
AUTH_TIMEOUT_SECONDS = 10
# The CSV / xlsx export endpoints stream a whole tab or workbook, so they get
# more room than an API call — but still a fixed ceiling.
EXPORT_TIMEOUT_SECONDS = 60


class SheetsUnavailable(RuntimeError):
    """A Sheets read could not be completed.

    Deliberately distinct from an *empty* result. "The API refused us" and "the
    workbook genuinely has no roll-up tab" are different facts, and collapsing
    them into ``{}`` is how a 429 turned into a silent, permanent-looking "this
    workbook has no official totals". Callers may present this to a user; the
    message always names the actual cause.
    """


class _TimedRequest:
    """google-auth transport wrapper that stamps our deadline onto token calls.

    ``creds.refresh(request)`` never passes a timeout, so google-auth's 120s
    default applies. Wrapping the transport is the only seam that reaches it.
    """

    def __init__(self, inner, timeout: float):
        self._inner = inner
        self._timeout = timeout

    def __call__(self, url, method="GET", body=None, headers=None, timeout=None, **kwargs):
        return self._inner(
            url, method=method, body=body, headers=headers,
            timeout=self._timeout if timeout is None else timeout, **kwargs,
        )


# ADC resolution is not free — it reads the key file or queries the Cloud Run
# metadata server — and `_default_fetcher` used to redo it for *every tab*, so
# one 6-tab ingest paid for it six times. Credentials are safe to share: the
# httplib2 transport built on top of them is what is per-client (below).
_creds_cache: dict[tuple[str, ...], object] = {}
_creds_lock = threading.Lock()
_refresh_lock = threading.Lock()


def cached_credentials(scopes: Sequence[str]):
    """ADC credentials for a scope set, resolved once per process."""
    key = tuple(sorted(scopes))
    creds = _creds_cache.get(key)
    if creds is not None:
        return creds
    import google.auth

    with _creds_lock:
        creds = _creds_cache.get(key)
        if creds is None:
            creds, _ = google.auth.default(scopes=list(key))
            _creds_cache[key] = creds
    return creds


def refresh_if_stale(creds) -> None:
    """Mint a token only when the cached one is missing or expired."""
    if getattr(creds, "valid", False):
        return
    from google.auth.transport.requests import Request

    with _refresh_lock:
        if getattr(creds, "valid", False):  # another thread beat us to it
            return
        creds.refresh(_TimedRequest(Request(), AUTH_TIMEOUT_SECONDS))


def timed_http(creds, timeout: float):
    """An authorised httplib2 transport carrying an explicit socket timeout.

    ``build(credentials=...)`` constructs its own transport internally, so
    passing ``http=`` is the only way to state a deadline. A *fresh*
    ``httplib2.Http`` per client is deliberate and not an oversight: httplib2
    is not thread-safe, and these clients are used from FastAPI worker threads.
    Building one costs nothing (the discovery document is bundled statically in
    google-api-python-client, so `build` opens no socket).
    """
    import google_auth_httplib2
    import httplib2

    return google_auth_httplib2.AuthorizedHttp(creds, http=httplib2.Http(timeout=timeout))


#: HTTP statuses worth another attempt — Google's own back-pressure and its
#: transient 5xx. A 403 or 404 is an answer, not a blip, and retrying it just
#: spends the cron's budget arriving at the same place.
_RETRYABLE_STATUS = frozenset({429, 500, 502, 503, 504})

#: Attempts per call, including the first. Three is chosen against the caller,
#: not in the abstract: the MR cron fires every 3 minutes, so the worst case
#: here (3 x SHEETS_TIMEOUT_SECONDS plus backoff, ~95s) still finishes before
#: the next fire, and an overlapping fire is caught by SheetPullBusy anyway.
_RETRY_ATTEMPTS = 3

#: Waits *between* attempts, so len() is _RETRY_ATTEMPTS - 1. Short on purpose:
#: a socket timeout has already spent 30s, and the thing being ridden out is a
#: blip rather than an outage.
_RETRY_BACKOFF_SECONDS = (1.0, 3.0)


def _is_transient(exc: BaseException) -> bool:
    """Whether *exc* is the kind of failure a second attempt might survive.

    ``TimeoutError`` covers ``socket.timeout``, which is what httplib2 raises
    when the deadline in :func:`timed_http` expires — and it is what the MR cron
    actually hits ("workbook unreadable: The read operation timed out", roughly
    4% of fires). ``OSError`` catches the connection resets and DNS blips
    underneath it; ``TimeoutError`` is already an ``OSError``, but naming it
    keeps the intent legible.
    """
    from googleapiclient.errors import HttpError

    if isinstance(exc, HttpError):
        return getattr(exc.resp, "status", None) in _RETRYABLE_STATUS
    return isinstance(exc, (TimeoutError, OSError))


def _execute(request, *, what: str):
    """``request.execute()`` that rides out a transient failure.

    The Sheets reads behind the Marketing Research cron are two calls — a
    ``spreadsheets().get`` for the tab list and one ``values().batchGet`` across
    every tab — and the batch one is big enough to occasionally overrun its 30s
    deadline. There was no retry at all, so a single blip failed the whole fire
    and the snapshot for that tick simply did not happen.

    Retries are deliberately *not* pushed down into ``num_retries`` on
    ``execute()``: that handles retryable HTTP statuses but not the socket
    timeout raised by the transport, which is the failure actually being seen.

    A non-transient error is re-raised immediately and unchanged — callers
    distinguish "the API refused us" from "the workbook is empty" and that
    distinction must survive this wrapper.
    """
    import time

    last: BaseException | None = None
    for attempt in range(_RETRY_ATTEMPTS):
        try:
            return request.execute()
        except Exception as exc:  # noqa: BLE001 — re-raised below unless transient
            if not _is_transient(exc):
                raise
            last = exc
            if attempt < len(_RETRY_BACKOFF_SECONDS):
                logger.warning(
                    "Sheets %s failed (%s); retrying in %ss (attempt %d/%d)",
                    what, exc, _RETRY_BACKOFF_SECONDS[attempt],
                    attempt + 2, _RETRY_ATTEMPTS,
                )
                time.sleep(_RETRY_BACKOFF_SECONDS[attempt])
    raise SheetsUnavailable(
        f"Sheets {what} failed after {_RETRY_ATTEMPTS} attempts: {last}"
    ) from last


def _sheets_service():
    from googleapiclient.discovery import build

    creds = cached_credentials([SHEETS_SCOPE, DRIVE_SCOPE])
    return build(
        "sheets", "v4",
        http=timed_http(creds, SHEETS_TIMEOUT_SECONDS),
        cache_discovery=False,
    )


def workbook_meta(spreadsheet_id: str, *, service=None) -> dict:
    """Workbook title + tab list in one Sheets API call:
    ``{"title", "tabs": [{gid, title, hidden, rows, cols}]}``."""
    svc = service or _sheets_service()
    meta = _execute(svc.spreadsheets().get(spreadsheetId=spreadsheet_id),
                    what="workbook metadata")
    tabs = []
    for s in meta.get("sheets", []):
        p = s["properties"]
        grid = p.get("gridProperties", {})
        tabs.append({
            "gid": p["sheetId"],
            "title": p["title"],
            "hidden": p.get("hidden", False),
            "rows": grid.get("rowCount"),
            "cols": grid.get("columnCount"),
        })
    title = (meta.get("properties") or {}).get("title") or spreadsheet_id
    return {"title": title, "tabs": tabs}


def list_tabs(spreadsheet_id: str, *, service=None) -> list[dict]:
    """Every tab via the Sheets API: ``{gid, title, hidden, rows, cols}``."""
    return workbook_meta(spreadsheet_id, service=service)["tabs"]


def fetch_tab_values(spreadsheet_id: str, title: str, *, service=None) -> list[list[str]]:
    """Read a tab's displayed values via the Sheets API as a list-of-lists of
    strings (FORMATTED_VALUE keeps the $ / % / , formatting parse_tracker strips)."""
    svc = service or _sheets_service()
    resp = _execute(
        svc.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=f"'{title}'", valueRenderOption="FORMATTED_VALUE"),
        what=f"tab read {title!r}",
    )
    return [["" if c is None else str(c) for c in row] for row in resp.get("values", [])]


def fetch_all_tab_values(spreadsheet_id: str, titles: list[str], *, service=None) -> dict[str, list[list[str]]]:
    """Read many tabs in one Sheets API call (values.batchGet)."""
    svc = service or _sheets_service()
    resp = _execute(
        svc.spreadsheets()
        .values()
        .batchGet(spreadsheetId=spreadsheet_id, ranges=[f"'{t}'" for t in titles], valueRenderOption="FORMATTED_VALUE"),
        what=f"batch read of {len(titles)} tabs",
    )
    out: dict[str, list[list[str]]] = {}
    for title, vr in zip(titles, resp.get("valueRanges", [])):
        out[title] = [["" if c is None else str(c) for c in row] for row in vr.get("values", [])]
    return out


def _offline() -> bool:
    """The MR agent's offline flag, read the same way as everywhere else."""
    return os.environ.get("MR_OFFLINE") == "1"


def _default_fetcher(spreadsheet_id: str, gid: str) -> str:
    # Credentials are resolved once per process and the token is reused until it
    # actually expires — this used to re-run ADC + mint a fresh token per tab.
    import httpx

    # The offline flag has to be enforced HERE, not by every caller remembering
    # to inject a fetcher: this is the export fallback, so stubbing the Sheets
    # API seam left it wide open and a test run pulled the live workbook twice.
    if _offline():
        raise CredentialMissingError("offline mode")

    creds = cached_credentials([DRIVE_SCOPE])
    refresh_if_stale(creds)
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
    resp = httpx.get(
        url, headers={"Authorization": f"Bearer {creds.token}"},
        follow_redirects=True, timeout=EXPORT_TIMEOUT_SECONDS,
    )
    resp.raise_for_status()
    return resp.text


def _default_xlsx_fetcher(spreadsheet_id: str) -> bytes:
    import httpx

    if _offline():
        raise CredentialMissingError("offline mode")

    creds = cached_credentials([DRIVE_SCOPE])
    refresh_if_stale(creds)
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    resp = httpx.get(
        url, headers={"Authorization": f"Bearer {creds.token}"},
        follow_redirects=True, timeout=EXPORT_TIMEOUT_SECONDS,
    )
    resp.raise_for_status()
    return resp.content


def is_rollup_title(title: str) -> bool:
    """True when the TAB NAME says roll-up (e.g. "Marketing 2026 Overall
    Report"). This is the workbook's layout boundary — vendors before it, ops
    sheets after — and it is stable whatever the dropdown leaves in A1."""
    return "overall" in (title or "").lower()


def is_rollup_scope(rows: list[list[str]]) -> bool:
    """True when A1 says the tab is currently showing the consolidated view.

    A1 is a dropdown: any tab can be left scoped to "All". That makes it a fine
    signal for "do not ingest THIS tab as a vendor" and a terrible one for
    "stop reading the workbook" — one mis-set dropdown on the first tab used to
    end the scan before a single vendor was read, and the pull still reported
    success. Only :func:`is_rollup_title` may stop the scan.
    """
    a1 = (rows[0][0] if rows and rows[0] else "").strip().lower()
    return a1 in ("all", "overall")


def is_rollup_tab(title: str, rows: list[list[str]]) -> bool:
    """True for the consolidated roll-up tab by either signal. Its numbers are
    the sum of the vendor tabs, so ingesting it alongside them double-counts
    every dollar."""
    return is_rollup_title(title) or is_rollup_scope(rows)


def is_rollup_platform(platform: str) -> bool:
    """True when a dataset-run platform key ("sheets:<tab>") points at the
    roll-up tab. Read-path guard: runs ingested before the fetch-time skip
    existed stay in storage forever and would double-count every dollar."""
    plat = str(platform or "")
    title = plat.split(":", 1)[1] if ":" in plat else plat
    return is_rollup_title(title)


# Two labels the roll-up tab really does repeat across its Projected — Not
# Actualized and Paying blocks. Named once so the two entries reading them can
# never drift apart.
#
# The revenue-amount row is deliberately NOT here: the live tab spells it
# "revenue amount sold ($) (not actualized)" in the Projected block and
# "revenue amount sold (not actualized)" in the Paying block, two distinct
# labels with one hit each. Carrying both spellings as candidates of one entry
# was a real production defect — occurrence counts per candidate, never across
# the tuple, so occurrence 2 found nothing and BOTH revenue figures (and the
# ROAS derived from one of them) dropped out of the payload entirely.
_DUP_SERVICES_SOLD_NA = ("total services sold (not actualized)",)
_DUP_AMOUNT_SOLD_NO_FEE_NA = ("revenue amount sold w/o setup fee (not actualized)",)

# Block headers inside the team block: each opens a section, is unique on the
# tab, and is what actually tells the repeated rows apart. A field disambiguated
# by an anchor is read ONLY from that anchor's window — the rows between it and
# the next anchor — so a stray copy of the label elsewhere on the tab can never
# be picked up, and deleting one block leaves the other block reading correctly
# instead of silently shifting under it.
#
# This replaces occurrence counting for the repeated rows, on live evidence:
# "revenue amount sold w/o setup fee (not actualized)" appears THREE times on
# the tab (Projected, Paying, and a blank stray down in the Goal & Blended
# section). Occurrence 2 is the Paying copy only for as long as all three stay
# put; delete the Projected copy and occurrence 2 lands on the blank stray while
# occurrence 1 serves the Paying figure under the Projected name.
_ANCHOR_PROJECTED_NA = "number of projected new clients (not actualized)"
_ANCHOR_PAYING_NA = "number of paying new clients (not actualized)"
# Boundary only — no field reads from it, but it closes the Paying window so the
# stray copy further down the tab falls outside every window.
_ANCHOR_INBOUND = "number of paying new clients (inbound sales pipeline)"
_OFFICIAL_BLOCK_ANCHORS = (_ANCHOR_PROJECTED_NA, _ANCHOR_PAYING_NA, _ANCHOR_INBOUND)

# Team-level rows the roll-up tab reports itself. These are THE official
# figures (the roll-up aggregates ledger/raw sources — Referral, Websites, … —
# that have no vendor tab, so a vendor-tab sum can NEVER reproduce them).
#
# ``field: (labels, disambiguator)``. Order within ``labels`` = priority, exact
# stripped-lowercase match. The disambiguator says which row wins when the label
# is not unique on the tab, and is one of two things:
#
#   int  — a 1-based occurrence within the block, the convention and tuple
#          position ``_TEAM_MAP`` / ``_canonical_block`` in snapshots.py use
#          (whose label strings these reuse — copied deliberately rather than
#          imported, because that module's routes are WORKSPACE_SHARED and
#          importing it here would pull its scoping into the report path).
#   str  — a block anchor from _OFFICIAL_BLOCK_ANCHORS: read this row only from
#          the section that anchor opens. Positional, so it survives a stray
#          copy of the label elsewhere on the tab.
#
# Neither is cosmetic: the Paying block repeats the Projected — Not Actualized
# block's rows, so first-match-wins quietly published the Projected figures
# under the Paying names. One block's numbers appearing under another block's
# rows is exactly the 2026-08-15 "wrong numbers" defect.
_OFFICIAL_FIELD_LABELS: dict[str, tuple[Sequence[str], int | str]] = {
    # --- investment ---------------------------------------------------------
    "budget": (["budget"], 1),
    "spend": (["spend"], 1),
    # Performance is blank for this row on the live sheet — the Investment
    # fallback in parse_official_totals is what resolves it when it is billed at
    # all. Both cells have been empty since June 2026, so this field is legitimately
    # ABSENT for recent months; it must never be reported as a $0 fee.
    "management_fees": (["management fees"], 1),
    # --- leads --------------------------------------------------------------
    "leads": (["leads"], 1),
    "qualified_leads": (["qualified leads"], 1),
    "qualified_lead_ratio_pct": (["qualified lead ratio"], 1),
    "lost_dnc_bad_lead": (["lost dnc (bad lead)"], 1),
    "cost_per_lead": (["cost per lead"], 1),
    "cost_per_qualified_lead": (["cost per qualified lead"], 1),
    # --- demos --------------------------------------------------------------
    "demos_booked": (["total demos booked (sdr+vapi+direct)", "total demos booked"], 1),
    "qual_demos_booked": (["qualified demos booked (sdr+vapi+direct)", "qualified demos booked"], 1),
    "demos_completed": (["demos completed (sdr+vapi+direct)", "total demos completed"], 1),
    # The Direct-only count, and the reason it is here: the report's Conversion
    # Rate is revenue clients / demos completed (DIRECT), not / the all-in count
    # above. Q1 48/248 = 19.35% and Q2 57/212 = 26.89% land on the published
    # figures exactly, where 48/272 = 17.65% does not. Without this row the two
    # cards sit next to each other and cannot be reconciled by the reader.
    "demos_completed_direct": (["total demos completed (direct)"], 1),
    "show_up_rate_pct": (["total show up rate (%) (sdr+vapi+direct)"], 1),
    "leads_to_demo_booked_pct": (["leads to demo booked overall"], 1),
    "leads_to_qual_demo_booked_pct": (["leads to qualified demo booked"], 1),
    "cost_per_demo_booked": (["cost per demo booked (sdr+vapi+direct)"], 1),
    "cost_per_qual_demo_booked": (["cost per qualified demo booked (sdr+vapi+direct)"], 1),
    "cost_per_demo_completed": (["cost per demo completed (sdr+vapi+direct)"], 1),
    # --- projected revenue (actualized) -------------------------------------
    "projected_new_clients": (["number of projected new clients (actualized)"], 1),
    "projected_services_sold": (["total projected services sold (actualized)"], 1),
    "projected_amount_sold": (["projected total amount sold ($) actualized",
                               "projected total amount sold ($) (actualized)"], 1),
    "projected_mrr_without_setup_fee": (
        ["projected mrr from new sales w/o set up fees (actualized)"], 1),
    # --- actualized revenue -------------------------------------------------
    "revenue_clients": (["number of revenue clients (actualized)"], 1),
    "services_sold": (["total services sold (actualized)"], 1),
    "revenue_amount_sold": (["revenue amount sold (actualized)"], 1),
    "revenue_amount_sold_without_setup_fee": (
        ["revenue amount sold w/o setup fee (actualized)"], 1),
    # --- projected, NOT actualized (read from the Projected block only) ------
    # The anchor rows themselves are unique labels and need no disambiguation.
    "projected_new_clients_not_actualized": ([_ANCHOR_PROJECTED_NA], 1),
    "services_sold_not_actualized": (_DUP_SERVICES_SOLD_NA, _ANCHOR_PROJECTED_NA),
    # Distinct label from the Paying block's, per the live tab — anchored anyway,
    # so the day someone normalises the two spellings nothing silently collides.
    "revenue_amount_sold_not_actualized": (
        ["revenue amount sold ($) (not actualized)"], _ANCHOR_PROJECTED_NA),
    "revenue_amount_sold_without_setup_fee_not_actualized": (
        _DUP_AMOUNT_SOLD_NO_FEE_NA, _ANCHOR_PROJECTED_NA),
    # --- paying, NOT actualized (read from the Paying block only) ------------
    "paying_new_clients": ([_ANCHOR_PAYING_NA], 1),
    "paying_services_sold": (_DUP_SERVICES_SOLD_NA, _ANCHOR_PAYING_NA),
    "paying_revenue_amount_sold": (
        ["revenue amount sold (not actualized)"], _ANCHOR_PAYING_NA),
    "paying_revenue_amount_sold_without_setup_fee": (
        _DUP_AMOUNT_SOLD_NO_FEE_NA, _ANCHOR_PAYING_NA),
    # --- inbound sales pipeline ---------------------------------------------
    "inbound_pipeline_revenue_amount_sold": (
        ["revenue amount sold (inbound sales pipeline)"], 1),
    # --- goal & blended financials ------------------------------------------
    # Reported against the NOT-actualized revenue: 452,607.70 / 480,000 = 94.29%
    # and 660,703.09 / 535,000 = 123.50%, both matching the published figures.
    # A ratio, so a multi-month roll-up must recompute it from the summed
    # components - never average it, and never sum it.
    "revenue_target_pct": (["percentage of revenue target goal",
                            "% of revenue target goal (not actualized)",
                            "percentage of revenue target goal (not actualized)"], 1),
    # The goal MOVES between periods (480,000 -> 535,000), so it is read from
    # the sheet every time and can never be pinned as a constant anywhere.
    "revenue_sold_goal": (["revenue sold goal amount"], 1),
    "average_deal_amount": (["average deal amount"], 1),
    # Ratios both: recompute from summed components when rolling months up.
    "conversion_rate_pct": (["conversion rate (%)"], 1),
    # The sheet's ROAS row is the ACTUALIZED one - 262,947.70 / 239,581.57 =
    # 109.75%, the published "ROAS - Actualized (%)". The not-actualized
    # sibling has no row at all and is derived below.
    "roas_pct": (["roas"], 1),
    "cac": (["cac"], 1),
}


def _expected_occurrences(
    field_map: dict[str, tuple[Sequence[str], int | str]],
) -> dict[tuple[str, ...], int]:
    """Labels the map claims at MORE than one occurrence, and the highest one.

    Currently empty: every repeated row in the map is anchored to its block
    instead. It stays because an int disambiguator is still supported, and a
    future entry that counts occurrences must not silently lose this guard.

    Occurrence matching has a failure mode of its own: if the sheet loses one of
    the two duplicate blocks, the survivor becomes occurrence 1 and would be
    published under the OTHER block's name — a Paying figure reported as
    Projected. Nothing in the row tells us which block survived, so a block that
    carries fewer copies than the map declares means every field sharing that
    label is reported absent. Absent is recoverable; a plausible wrong number on
    a board report is not.
    """
    out: dict[tuple[str, ...], int] = {}
    for labels, where in field_map.values():
        if not isinstance(where, int):
            continue  # anchored fields are bounded by their block, not counted
        key = tuple(labels)
        out[key] = max(out.get(key, 0), where)
    return {labels: occ for labels, occ in out.items() if occ > 1}


_EXPECTED_OCCURRENCES = _expected_occurrences(_OFFICIAL_FIELD_LABELS)


def _anchor_windows(rows: list[list[str]], start: int, end: int) -> dict[str, tuple[int, int]]:
    """``anchor label -> (first row after it, first row of the next section)``.

    Only anchors actually present are returned, so a field whose anchor has been
    renamed away resolves to nothing and is reported absent — the same rule the
    parsed rows follow. Guessing which section a repeated row belonged to is the
    one outcome worth avoiding.
    """
    found = []
    for anchor in _OFFICIAL_BLOCK_ANCHORS:
        i = _row_for(rows, start, end, [anchor])
        if i is not None:
            found.append((i, anchor))
    found.sort()
    bounds = [i for i, _ in found] + [end]
    return {anchor: (i + 1, bounds[n + 1]) for n, (i, anchor) in enumerate(found)}


def _derive_official_fields(fields: dict[str, float]) -> None:
    """Add the official figures that have no row of their own, in place.

    Derived from the components of the SAME month column, never from an average
    across months: averaging the monthly ROAS figures instead of recomputing it
    from summed revenue and spend measures +1.74pp of error on this data. A
    caller that rolls several months up must therefore recompute this from the
    summed components, not sum or average this field.

    A derived field is emitted only when every input is present and usable, so a
    renamed or removed sheet row leaves it ABSENT — the same rule the parsed
    fields follow. A missing metric must never arrive as 0.
    """
    revenue = fields.get("revenue_amount_sold_not_actualized")
    spend = fields.get("spend")
    if revenue is None or not spend:
        return  # missing input, or a zero denominator — no honest ratio exists
    # Percent, matching the sheet's own ROAS row (a "188.92%" cell parses to
    # 188.92), so the derived figure and the parsed one are comparable.
    fields["roas_not_actualized_pct"] = round(revenue / spend * 100, 2)


def parse_official_totals(
    rows: list[list[str]], year: int, *, warnings: list[str] | None = None,
) -> dict[str, dict[str, float]]:
    """The roll-up tab's own team-level rows — the sheet's official monthly
    figures, the exact cells the team reads (Performance basis, Investment
    fallback). Keys are "YYYY-MM" → {field: value}.

    ``warnings`` (optional, appended to) collects layout notes. The roll-up is
    the one tab whose figures nothing else can cross-check field by field, so a
    repeated month band here needs saying even when the numbers look plausible.
    """
    if not rows:
        return {}
    months, repeated = scan_month_columns(rows[0])
    if not months:
        return {}
    title = (rows[0][0] if rows[0] else "").strip()
    if repeated and warnings is not None:
        warnings.append(_repeat_gap_message(title or "Overall tab", repeated))
    blocks = _find_blocks(rows, title)
    start, end = (blocks[0][1], blocks[0][2]) if blocks else (1, len(rows))
    out: dict[str, dict[str, float]] = {}
    windows = _anchor_windows(rows, start, end)
    for field, (labels, where) in _OFFICIAL_FIELD_LABELS.items():
        if isinstance(where, str):
            window = windows.get(where)
            if window is None:
                continue  # the block header that identifies this row is gone
            lo, hi = window
            i = _row_for(rows, lo, hi, labels)
            if i is not None and _row_for(rows, lo, hi, labels, 2) is not None:
                continue  # two copies inside one section - ambiguous, withhold
        else:
            expected = _EXPECTED_OCCURRENCES.get(tuple(labels))
            if expected and _row_for(rows, start, end, labels, expected) is None:
                continue  # a duplicate block went missing - report none, guess none
            i = _row_for(rows, start, end, labels, where)
        if i is None:
            continue  # renamed/removed row -> absent field, never a zero
        row = rows[i]
        cell = lambda c: row[c] if 0 <= c < len(row) else ""
        for month, perf, inv in months:
            v = _num(cell(perf))
            if v is None and inv >= 0:
                v = _num(cell(inv))
            if v is not None:
                out.setdefault(f"{year:04d}-{month:02d}", {})[field] = v
    for fields in out.values():
        _derive_official_fields(fields)
    return out


def parse_official_spend(rows: list[list[str]], year: int) -> dict[str, float]:
    """Spend-only view of :func:`parse_official_totals` (kept for callers that
    predate the full-totals read)."""
    return {k: v["spend"] for k, v in parse_official_totals(rows, year).items() if "spend" in v}


def fetch_official_totals(
    spreadsheet_id: str, year: int, *, service=None, warnings: list[str] | None = None,
) -> dict[str, dict[str, float]]:
    """Official monthly team-level figures from the consolidated roll-up tab.

    Returns ``{}`` for exactly one meaning: **the workbook was read successfully
    and has no roll-up tab carrying official rows.** Any failure to read it — a
    429, a revoked share, a timeout, the Sheets API disabled — raises
    :class:`SheetsUnavailable` instead.

    This used to ``return {}`` on any exception, which made a transient Sheets
    error indistinguishable from a workbook that genuinely has no Overall tab:
    the headline figures silently vanished and nothing anywhere said why. The
    caller needs to tell those apart, because they warrant opposite responses —
    surface an error vs. label the figure as a vendor-tab sum.
    """
    try:
        svc = service or _sheets_service()
        tabs = list_tabs(spreadsheet_id, service=svc)
    except Exception as exc:  # noqa: BLE001 — re-raised, never swallowed
        raise SheetsUnavailable(
            f"could not list the tabs of workbook {spreadsheet_id}: {exc}"
        ) from exc

    for tab in tabs:
        if not is_rollup_title(tab["title"]):
            continue
        try:
            # 120 rows of headroom over a tab whose team block is 81 labelled
            # rows today. What falls off first is the TAIL - the Goal & Blended
            # section (% of revenue target goal, revenue sold goal, average deal
            # amount, conversion rate, ROAS, CAC), CAC being the 79th label - so
            # rows inserted higher up cost the KPIs, not the funnel. Raise this
            # before the tab grows, not after the board report loses its ratios.
            rows = fetch_tab_values(spreadsheet_id, tab["title"], service=svc)[:120]
        except Exception as exc:  # noqa: BLE001 — re-raised, never swallowed
            raise SheetsUnavailable(
                f'could not read the roll-up tab "{tab["title"]}" of workbook '
                f"{spreadsheet_id}: {exc}"
            ) from exc
        official = parse_official_totals(rows, year, warnings=warnings)
        if official:
            return official
    # Read fine, nothing to report — the honest empty case.
    return {}


def fetch_official_spend(spreadsheet_id: str, year: int, *, service=None) -> dict[str, float]:
    """Spend-only view of :func:`fetch_official_totals` (back-compat).

    Raises :class:`SheetsUnavailable` on a read failure, same contract.
    """
    totals = fetch_official_totals(spreadsheet_id, year, service=service)
    return {k: v["spend"] for k, v in totals.items() if "spend" in v}


# The roll-up aggregates the vendor tabs PLUS ledger sources only it carries
# (Referral, Websites, …). Its media spend is therefore always >= the sum of the
# vendor tabs. When it comes back BELOW them, that is not a business fact — it is
# a read that landed on the wrong cells, and it is the single check that would
# have caught the 2026-08 breakage on the day it happened instead of a month
# later. A little slack absorbs rounding and the seconds between two reads.
OFFICIAL_SHORTFALL_TOLERANCE = 0.98


def reconcile_official_spend(
    fetched: list[dict], official: dict[str, dict[str, float]], *, through: date
) -> list[str]:
    """Months whose official spend is impossibly below the vendor-tab sum.

    Compared on media channels only (the vendor tabs' non-media blocks are the
    part we cannot be sure the roll-up's Spend row includes) and only for months
    up to ``through`` — vendor tabs pre-fill future retainer months that the
    roll-up legitimately leaves at zero.

    An empty list means the two halves agree. Every entry names both figures, so
    whoever reads it can go straight to the cell.
    """
    by_month: dict[str, float] = {}
    for tab in fetched:
        for m in tab.get("metrics") or []:
            if m.channel in config.NON_MEDIA_CHANNELS:
                continue
            key = f"{m.date.year:04d}-{m.date.month:02d}"
            by_month[key] = by_month.get(key, 0.0) + float(m.spend or 0.0)
    limit = f"{through.year:04d}-{through.month:02d}"
    out: list[str] = []
    for key, vendor_sum in sorted(by_month.items()):
        if key > limit or vendor_sum <= 0:
            continue
        off = (official.get(key) or {}).get("spend")
        if off is None:
            continue
        if off < vendor_sum * OFFICIAL_SHORTFALL_TOLERANCE:
            out.append(
                f"{key}: the Overall tab reports ${off:,.2f} spend but the vendor "
                f"tabs sum to ${vendor_sum:,.2f} — the roll-up cannot be below "
                f"the tabs it aggregates, so its columns are being misread"
            )
    return out


def fetch_all_trackers(
    spreadsheet_id: str,
    year: int,
    *,
    service=None,
    xlsx_fetcher: Callable[[str], bytes] | None = None,
    max_rows: int = 200,
) -> list[dict]:
    """Discover and parse every tab that is a performance tracker.

    Prefers the Google Sheets API (clean tab enumeration with gids/titles). Falls
    back to a whole-workbook xlsx export (openpyxl) if the Sheets API is
    unavailable. Non-tracker tabs (Looker dumps, raw data, pivots) parse to
    nothing and are skipped. Returns ``{"tab", "gid"?, "metrics", "gaps"}``.

    An empty list means "read fine, no tab parsed as a tracker". If *both* read
    paths fail this raises :class:`SheetsUnavailable` naming both causes — the
    API error is the diagnostic one, and reporting only the fallback's failure
    sends whoever is debugging down the wrong road.
    """
    try:
        svc = service or _sheets_service()
        out: list[dict] = []
        for tab in list_tabs(spreadsheet_id, service=svc):
            # Hidden tabs are archives and Looker dumps, never a live vendor —
            # and they sit BEFORE the vendors, so this also saves their fetches.
            if tab.get("hidden"):
                continue
            rows = fetch_tab_values(spreadsheet_id, tab["title"], service=svc)[:max_rows]
            if is_rollup_title(tab["title"]):
                # Workbook layout rule (user, 2026-07-27): every tab BEFORE the
                # Overall Report is a vendor; everything after it is ops sheets
                # (raw data, HubSpot pulls, logs) — never ingest past it.
                break
            if is_rollup_scope(rows):
                continue  # left on the "All" scope — skip the tab, keep scanning
            metrics, gaps = parse_tracker(rows, year)
            if metrics:
                out.append({"tab": tab["title"], "gid": tab["gid"], "metrics": metrics, "gaps": gaps})
        return out
    except Exception as api_exc:  # noqa: BLE001 — the xlsx path is a real fallback
        try:
            return _fetch_all_trackers_xlsx(
                spreadsheet_id, year, xlsx_fetcher=xlsx_fetcher, max_rows=max_rows
            )
        except Exception as xlsx_exc:  # noqa: BLE001 — both paths down = honest raise
            raise SheetsUnavailable(
                f"could not read workbook {spreadsheet_id}: the Sheets API failed "
                f"({api_exc}) and the xlsx export fallback also failed ({xlsx_exc})"
            ) from api_exc


def _fetch_all_trackers_xlsx(
    spreadsheet_id: str,
    year: int,
    *,
    xlsx_fetcher: Callable[[str], bytes] | None = None,
    max_rows: int = 200,
) -> list[dict]:
    """Fallback discovery: export the whole workbook as xlsx and scan every tab.
    Requires ``openpyxl`` (imported lazily)."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError(
            "openpyxl is required for the xlsx discovery fallback; "
            "`pip install openpyxl`, enable the Sheets API, or pass an explicit gid."
        ) from exc
    data = (xlsx_fetcher or _default_xlsx_fetcher)(spreadsheet_id)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict] = []
    for name in wb.sheetnames:
        ws = wb[name]
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue  # same rule as the API path
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
            if len(rows) >= max_rows:
                break
        if is_rollup_title(name):
            break  # same layout rule as the API path: vendors end at the roll-up
        if is_rollup_scope(rows):
            continue
        metrics, gaps = parse_tracker(rows, year)
        if metrics:
            out.append({"tab": name, "metrics": metrics, "gaps": gaps})
    return out


class SheetsSource:
    def __init__(
        self,
        spreadsheet_id: str,
        gid: str,
        *,
        year: int,
        brand: str | None = None,
        fetcher: Callable[[str, str], str] | None = None,
    ):
        self.spreadsheet_id = spreadsheet_id
        self.gid = str(gid)
        self.year = year
        self.brand = brand
        self.name = f"sheets:{spreadsheet_id}:{gid}"
        self._fetcher = fetcher or _default_fetcher

    def _rows(self) -> list[list[str]]:
        text = self._fetcher(self.spreadsheet_id, self.gid)
        return list(csv.reader(io.StringIO(text)))

    def fetch_campaign_metrics(self, range: DateRange) -> tuple[list[CampaignMetric], list[DataGap]]:
        metrics, gaps = parse_tracker(self._rows(), self.year, self.brand)
        lo, hi = range.start, range.end
        kept = [m for m in metrics if lo <= m.date <= hi]
        return kept, gaps

    def fetch_leads(self, range: DateRange) -> tuple[list[Lead], list[DataGap]]:
        # This tracker is channel-aggregate; it carries no lead-level rows.
        return [], []
