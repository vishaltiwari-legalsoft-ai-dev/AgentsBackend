import csv
import io
import os
from datetime import date

from marketing_research_agent.schemas import DateRange
from marketing_research_agent.sources.sheets_source import SheetsSource, parse_tracker

FIX = os.path.join(os.path.dirname(__file__), "fixtures", "tracker_tab.csv")
RANGE = DateRange(start=date(2026, 1, 1), end=date(2026, 12, 31))


def _rows():
    with open(FIX, newline="", encoding="utf-8") as fh:
        return list(csv.reader(fh))


def test_parse_splits_channel_blocks():
    metrics, gaps = parse_tracker(_rows(), year=2026)
    channels = {m.channel for m in metrics}
    assert channels == {"META", "Google"}
    # two months of data per channel
    assert len(metrics) == 4


def test_meta_spend_uses_performance_column():
    """Official basis (decision 2026-07-27): the console must show the figures
    the team reads on the sheet — the Performance column, not Investment."""
    metrics, _ = parse_tracker(_rows(), year=2026)
    jan_meta = next(m for m in metrics if m.channel == "META" and m.date == date(2026, 1, 1))
    assert jan_meta.spend == 3603.88  # Performance, not Investment (3964.27)
    assert jan_meta.leads == 27 and jan_meta.qualified_leads == 21
    assert jan_meta.demos_booked == 11 and jan_meta.demos_completed == 4


def test_spend_falls_back_to_investment_when_performance_blank():
    rows = [
        ["Meta TestBrand", "Jan (Performance)", "Jan (Investment)"],
        ["Spend", "", "$500.00"],
        ["Leads", "3", ""],
    ]
    metrics, _ = parse_tracker(rows, year=2026)
    assert metrics[0].spend == 500.0


def test_brand_derived_from_title():
    metrics, _ = parse_tracker(_rows(), year=2026)
    assert all(m.utm_campaign == "TestBrand" for m in metrics)


def test_sheets_source_with_injected_fetcher():
    with open(FIX, encoding="utf-8") as fh:
        text = fh.read()
    src = SheetsSource("sheet1", "0", year=2026, fetcher=lambda sid, gid: text)
    metrics, gaps = src.fetch_campaign_metrics(RANGE)
    assert len(metrics) == 4


class _FakeValues:
    def __init__(self, by_title):
        self._by_title = by_title

    def get(self, spreadsheetId, range, valueRenderOption=None):
        title = range.strip("'")
        rows = self._by_title.get(title, [])
        return _FakeExec({"values": rows})


class _FakeSpreadsheets:
    def __init__(self, tabs, by_title):
        self._tabs = tabs
        self._values = _FakeValues(by_title)

    def get(self, spreadsheetId):
        return _FakeExec({
            "sheets": [
                {"properties": {"sheetId": t["gid"], "title": t["title"],
                                "hidden": t.get("hidden", False),
                                "gridProperties": {"rowCount": 100, "columnCount": 30}}}
                for t in self._tabs
            ]
        })

    def values(self):
        return self._values


class _FakeExec:
    def __init__(self, payload):
        self._payload = payload

    def execute(self):
        return self._payload


class _FakeService:
    def __init__(self, tabs, by_title):
        self._s = _FakeSpreadsheets(tabs, by_title)

    def spreadsheets(self):
        return self._s


def test_fetch_all_trackers_via_sheets_api_skips_non_trackers_and_rollups():
    """Primary discovery path (Sheets API): a vendor tracker + the consolidated
    roll-up + a junk tab -> only the vendor tracker is ingested (the roll-up
    duplicates the vendors' numbers and would double-count)."""
    from marketing_research_agent.sources.sheets_source import fetch_all_trackers

    tabs = [
        {"gid": 559258152, "title": "Meta 360 RA"},
        {"gid": 2088778899, "title": "Marketing 2026 Overall Report"},
        {"gid": 12345, "title": "Raw Notes"},
    ]
    by_title = {
        "Meta 360 RA": _rows(),
        "Marketing 2026 Overall Report": _rows(),  # parses fine; skipped by title
        "Raw Notes": [["just", "notes"], ["no", "months"]],
    }
    found = fetch_all_trackers("sheet1", 2026, service=_FakeService(tabs, by_title))
    assert len(found) == 1
    assert found[0]["tab"] == "Meta 360 RA" and found[0]["gid"] == 559258152
    assert len(found[0]["metrics"]) == 4


def test_parse_official_spend_reads_rollup_performance_spend():
    from marketing_research_agent.sources.sheets_source import parse_official_spend

    rows = [
        ["All", "Jan (Performance)", "Jan (Investment)", "1st Quarter",
         "July (Performance)", "July (Investment)"],
        ["Budget", "$100.00", "", "", "$77,100.20", "#N/A"],
        ["Spend", "$1,000.50", "#N/A", "", "$45,461.20", "#N/A"],
        ["Leads", "10", "", "", "305", ""],
    ]
    assert parse_official_spend(rows, 2026) == {"2026-01": 1000.50, "2026-07": 45461.20}


def test_fetch_official_spend_reads_the_rollup_tab():
    from marketing_research_agent.sources.sheets_source import fetch_official_spend

    tabs = [
        {"gid": 1, "title": "Meta 360 RA"},
        {"gid": 2, "title": "Marketing 2026 Overall Report"},
    ]
    by_title = {
        "Meta 360 RA": _rows(),
        "Marketing 2026 Overall Report": [
            ["All", "July (Performance)", "July (Investment)"],
            ["Spend", "$45,461.20", "#N/A"],
        ],
    }
    out = fetch_official_spend("sheet1", 2026, service=_FakeService(tabs, by_title))
    assert out == {"2026-07": 45461.20}


def _meta_tab_with_a_pasted_websites_block():
    """A vendor tab shaped like the live ones on 2026-08-15: its own META block,
    then a copy of the shared Websites block underneath."""
    return [
        ["Axenic LS Meta", "Aug (Performance)", "Aug (Investment)"],
        ["Spend", "$0.00", "#N/A"],
        ["Leads", "0", ""],
        ["Websites"],
        ["Spend", "$8,632.00", "$8,632.00"],
        ["Leads", "127", ""],
    ]


def test_a_websites_block_pasted_into_a_vendor_tab_is_not_ingested():
    """Regression, 2026-08-15: the shared Websites block sat in SEVEN vendor
    tabs plus its own. Ingesting it from a vendor tab attributed Websites'
    numbers to that vendor AND counted $8,632/month eight times."""
    metrics, gaps = parse_tracker(_meta_tab_with_a_pasted_websites_block(), year=2026)
    assert metrics == []                      # its own block is empty; nothing else is its
    assert any("pasted into this tab" in g.message for g in gaps)


def test_the_websites_tab_itself_still_reports_its_block():
    """The dedicated tab is the one place those rows are real."""
    rows = [
        ["Websites", "Aug (Performance)", "Aug (Investment)"],
        ["Spend", "$8,632.00", "$8,632.00"],
        ["Leads", "127", ""],
    ]
    metrics, _ = parse_tracker(rows, year=2026)
    assert [(m.channel, m.spend, m.leads) for m in metrics] == [("Websites", 8632.0, 127)]


def test_an_unknown_vendor_channel_is_other_not_total():
    """"Total" is the roll-up's channel, and _campaign_structured POPS it out as
    the report's totals. A vendor tab whose channel is unrecognised
    ("DanteAgency ChatGPT") therefore became the headline KPI strip."""
    rows = [["Barnacle Ads LS", "Aug (Performance)"], ["Spend", "$43.16"], ["Leads", "2"]]
    metrics, _ = parse_tracker(rows, year=2026)
    assert metrics[0].channel == "Other"

    # The consolidated roll-up (A1 = "All") keeps "Total" — that one IS the totals.
    rollup = [["All", "Aug (Performance)"], ["Spend", "$30,431.46"], ["Leads", "9"]]
    assert parse_tracker(rollup, year=2026)[0][0].channel == "Total"


def test_hidden_tabs_are_never_ingested():
    """Hidden tabs are Looker dumps and archives, and in the live workbook six of
    them sit BEFORE the vendors."""
    from marketing_research_agent.sources.sheets_source import fetch_all_trackers

    tabs = [
        {"gid": 1, "title": "Looker Studio per Brand (April)", "hidden": True},
        {"gid": 2, "title": "Meta 360 RA", "hidden": False},
    ]
    by_title = {"Looker Studio per Brand (April)": _rows(), "Meta 360 RA": _rows()}
    found = fetch_all_trackers("sheet1", 2026, service=_FakeService(tabs, by_title))
    assert [f["tab"] for f in found] == ["Meta 360 RA"]


def test_repeated_month_band_reads_the_leftmost_grid_and_says_so():
    """Regression, 2026-08: a second month-column band added to the right of the
    grid silently re-pointed every figure — the same Overall tab that read July
    spend $48,005.09 on 07-27 read $5,579.58 afterwards, a roll-up BELOW the
    vendor tabs it aggregates. The primary grid is the leftmost one."""
    from marketing_research_agent.sources.sheets_source import (
        parse_official_totals, scan_month_columns,
    )

    header = ["All", "July (Performance)", "July (Investment)", "1st Quarter",
              "YTD (Performance)", "July (Performance)", "July (Investment)"]
    cols, repeated = scan_month_columns(header)
    assert cols == [(7, 1, 2)]          # leftmost band, not columns 5/6
    assert repeated == [7]              # and the repeat is reported, not swallowed

    rows = [header, ["Spend", "$48,005.09", "", "", "", "$5,579.58", ""]]
    assert parse_official_totals(rows, 2026) == {"2026-07": {"spend": 48005.09}}


def test_investment_is_never_paired_across_bands():
    """Performance and Investment used to resolve independently, so a pair could
    be lifted from two different tables. No billed figure beats a wrong one."""
    from marketing_research_agent.sources.sheets_source import scan_month_columns

    # Investment sits LEFT of the Performance column that survived — different band.
    cols, _ = scan_month_columns(["All", "Jan (Investment)", "Feb (Performance)",
                                  "Jan (Performance)"])
    assert cols == [(1, 3, -1), (2, 2, -1)]


def test_parse_tracker_reports_a_repeated_month_as_a_gap():
    rows = [
        ["Meta TestBrand", "Jan (Performance)", "Jan (Investment)", "Jan (Performance)"],
        ["Spend", "$100.00", "", "$999.00"],
        ["Leads", "3", "", "99"],
    ]
    metrics, gaps = parse_tracker(rows, year=2026)
    assert metrics[0].spend == 100.0 and metrics[0].leads == 3
    assert any("more than one column band" in g.message for g in gaps)


def test_a_vendor_tab_left_on_the_all_scope_is_skipped_not_a_full_stop():
    """A1 is a dropdown — any tab can be left scoped to "All". That used to end
    the whole scan, so a pull ingested ZERO vendor tabs and still reported
    success. Only the tab NAME may stop the scan."""
    from marketing_research_agent.sources.sheets_source import fetch_all_trackers

    tabs = [
        {"gid": 1, "title": "Meta 360 RA"},          # dropdown left on "All"
        {"gid": 2, "title": "Hawksem LS Google"},    # a real vendor after it
        {"gid": 3, "title": "Marketing 2026 Overall Report"},
        {"gid": 4, "title": "Raw Notes"},
    ]
    scoped_to_all = [["All"] + _rows()[0][1:]] + _rows()[1:]
    by_title = {
        "Meta 360 RA": scoped_to_all,
        "Hawksem LS Google": _rows(),
        "Marketing 2026 Overall Report": _rows(),
        "Raw Notes": [["just", "notes"]],
    }
    found = fetch_all_trackers("sheet1", 2026, service=_FakeService(tabs, by_title))
    assert [f["tab"] for f in found] == ["Hawksem LS Google"]


def test_is_rollup_tab():
    from marketing_research_agent.sources.sheets_source import is_rollup_tab

    # title match survives whatever the dropdown left in A1
    assert is_rollup_tab("Marketing 2026 Overall Report", [["Meta 360 RA"]]) is True
    # A1 scope says it's the consolidated view
    assert is_rollup_tab("Some Tab", [["All", "Jan (Performance)"]]) is True
    assert is_rollup_tab("Some Tab", [["Overall"]]) is True
    # a plain vendor tab is not a roll-up
    assert is_rollup_tab("Meta 360 RA", [["Meta 360 RA"]]) is False
    assert is_rollup_tab("Meta 360 RA", []) is False


def _metric(channel, month, spend):
    from marketing_research_agent.schemas import CampaignMetric

    return CampaignMetric(channel=channel, campaign="c", utm_source="s", utm_medium="paid",
                          utm_campaign="b", spend=spend, leads=0, qualified_leads=0,
                          demos_booked=0, demos_completed=0, date=date(2026, month, 1))


def test_reconcile_flags_a_rollup_below_the_tabs_it_aggregates():
    """The check that would have caught 2026-08 on the day it happened: the
    Overall tab cannot report less media spend than the vendor tabs it sums."""
    from marketing_research_agent.sources.sheets_source import reconcile_official_spend

    fetched = [{"metrics": [_metric("META", 7, 30000.0), _metric("Google", 7, 9945.26)]}]
    official = {"2026-07": {"spend": 5579.58}}
    out = reconcile_official_spend(fetched, official, through=date(2026, 8, 15))
    assert len(out) == 1
    assert "$5,579.58" in out[0] and "$39,945.26" in out[0]


def test_reconcile_is_quiet_when_the_rollup_is_the_larger_figure():
    """The healthy shape — the roll-up adds ledger sources no vendor tab carries."""
    from marketing_research_agent.sources.sheets_source import reconcile_official_spend

    fetched = [{"metrics": [_metric("META", 7, 34873.09)]}]
    official = {"2026-07": {"spend": 48005.09}}
    assert reconcile_official_spend(fetched, official, through=date(2026, 8, 15)) == []


def test_reconcile_ignores_prefilled_future_months_and_non_media():
    """Vendor tabs pre-fill future retainer months the roll-up leaves at zero,
    and Websites spend is not media spend — neither is a mismatch."""
    from marketing_research_agent.sources.sheets_source import reconcile_official_spend

    fetched = [{"metrics": [_metric("META", 9, 10186.0), _metric("Websites", 7, 7486.0)]}]
    official = {"2026-07": {"spend": 100.0}, "2026-09": {"spend": 0.0}}
    assert reconcile_official_spend(fetched, official, through=date(2026, 8, 15)) == []


def test_xlsx_discovery_fallback_skips_rollups():
    """Fallback discovery path (whole-workbook xlsx via openpyxl)."""
    import io as _io

    import openpyxl

    from marketing_research_agent.sources.sheets_source import _fetch_all_trackers_xlsx

    wb = openpyxl.Workbook()
    tracker = wb.active
    tracker.title = "Meta 360 RA"
    for r in _rows():
        tracker.append(r)
    rollup = wb.create_sheet("Marketing 2026 Overall Report")
    for r in _rows():
        rollup.append(r)
    junk = wb.create_sheet("Raw Notes")
    junk.append(["just", "some", "notes"])
    buf = _io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    found = _fetch_all_trackers_xlsx("sheet1", 2026, xlsx_fetcher=lambda sid: data)
    assert len(found) == 1 and found[0]["tab"] == "Meta 360 RA"
# --- roll-up tab: the live label sequence + the report ledger ----------------
# _LIVE_LABELS is the roll-up tab's team block exactly as production carries it:
# 81 labelled rows dumped read-only from mr_snapshots/
# marketing-2026-overall-report_2026-09-04, byte-identical across the oldest,
# middle and newest of 59 stored snapshots, taken through the same
# _find_blocks(...)[0] path parse_official_totals uses.
#
# It is embedded rather than invented because an invented one hid a production
# defect behind a green gate: the fixture gave BOTH blocks the plain label
# "Revenue Amount Sold (Not Actualized)", manufacturing a duplicate the live tab
# does not have. The tab actually spells row 61 "Revenue Amount Sold ($) (Not
# Actualized)" (Projected) and row 65 "Revenue Amount Sold (Not Actualized)"
# (Paying) - two distinct labels, one hit each - so a single entry carrying both
# spellings as candidates resolved neither, and two revenue figures plus the
# ROAS derived from one of them were absent in production.
#
# Three properties of this sequence are load-bearing, all of them real:
#   rows 60 / 64  "Total Services Sold (Not Actualized)" genuinely twice
#   rows 62 / 66  "...w/o Setup Fee (Not Actualized)" twice, PLUS
#   row  74       a third, blank, stray copy down in Goal & Blended
#   rows 59 / 63  the unique block headers everything above is read relative to
_LIVE_LABELS = [
    'Management Fees',                                         #  1
    'Budget',                                                  #  2
    'Spend',                                                   #  3
    'Leads',                                                   #  4
    'Qualified Leads',                                         #  5
    'Qualified Lead Ratio',                                    #  6
    'Cost per Lead',                                           #  7
    'Cost per Qualified Lead',                                 #  8
    'Total Demos Booked (SDR+VAPI+Direct)',                    #  9
    'Total Demos Booked (Direct)',                             # 10
    'Leads to Demo Booked Overall',                            # 11
    'Leads to Qualified Demo Booked',                          # 12
    'SDR Demos Booked',                                        # 13
    'SDR Inbound Demo Booked',                                 # 14
    'SDR Rescheduled Demo Booked',                             # 15
    'SDR Inbound Bad Lead Helper',                             # 16
    'SDR Demos Completed',                                     # 17
    'SDR Resched Completed',                                   # 18
    'SDR Inbound Completed',                                   # 19
    'SDR Demo Completed Ratio',                                # 20
    'VAPI Demos Booked',                                       # 21
    'VAPI Demos Completed',                                    # 22
    'VAPI New Demo',                                           # 23
    'VAPI New Demos Completed',                                # 24
    'VAPI Resched',                                            # 25
    'VAPI Resched Demos Completed',                            # 26
    'VAPI Demo Completed Ratio',                               # 27
    'Qualified Demos Booked (Direct)',                         # 28
    'Qualified Demos Booked (SDR+VAPI+Direct)',                # 29
    'Qualified Demos Ratio over Total Demos',                  # 30
    'Total Demos Completed (Direct)',                          # 31
    'Hot Leads - Follow Up <90 days',                          # 32
    'Cold Stage 3 months',                                     # 33
    'Cold Stage 6 months',                                     # 34
    'Cold Stage 12 months',                                    # 35
    'No Show',                                                 # 36
    'Canceled',                                                # 37
    'Demos Completed (SDR+VAPI+Direct)',                       # 38
    'Total Show Up Rate (%) (SDR+VAPI+Direct)',                # 39
    'Total Show Up Rate (%) (Direct)',                         # 40
    'Qualified Lead to Demo Booked (%)',                       # 41
    'Lost DNC (Bad Lead)',                                     # 42
    'Not a valid lead (applicant)',                            # 43
    'Not Partnership-fit (Lead)',                              # 44
    'Wrong contact info',                                      # 45
    'Cost Per Qualified Demo Booked (Direct)',                 # 46
    'Cost Per Qualified Demo Booked (SDR+VAPI+Direct)',        # 47
    'Cost Per Demo Booked (SDR+VAPI+Direct)',                  # 48
    'Cost Per Demo Completed (Direct Demos)',                  # 49
    'Cost Per Demo Completed (SDR+VAPI+Direct)',               # 50
    'Number of Projected New Clients (Actualized)',            # 51
    'Total Projected Services Sold (Actualized)',              # 52
    'Projected Total Amount Sold ($) Actualized',              # 53
    'Projected MRR from New Sales w/o set up fees (Actualized)', # 54
    'Number of Revenue Clients (Actualized)',                  # 55
    'Total Services Sold (Actualized)',                        # 56
    'Revenue Amount Sold (Actualized)',                        # 57
    'Revenue Amount Sold w/o Setup Fee (Actualized)',          # 58
    'Number of Projected New Clients (Not Actualized)',        # 59
    'Total Services Sold (Not Actualized)',                    # 60
    'Revenue Amount Sold ($) (Not Actualized)',                # 61
    'Revenue Amount Sold w/o Setup Fee (Not Actualized)',      # 62
    'Number of Paying New Clients (Not Actualized)',           # 63
    'Total Services Sold (Not Actualized)',                    # 64
    'Revenue Amount Sold (Not Actualized)',                    # 65
    'Revenue Amount Sold w/o Setup Fee (Not Actualized)',      # 66
    'Number of Paying New Clients (Inbound Sales Pipeline)',   # 67
    'Total Services Sold (Inbound Sales Pipeline)',            # 68
    'Revenue Amount Sold (Inbound Sales Pipeline)',            # 69
    'Revenue Amount Sold w/o Setup Fee (Inbound Sales Pipeline)', # 70
    'Percentage of Revenue Target Goal',                       # 71
    'Revenue Sold Goal Amount',                                # 72
    'Revenue Amount Sold (Lead Financials)',                   # 73
    'Revenue Amount Sold w/o Setup Fee (Not Actualized)',      # 74
    'Confirmed All Revenue (MRR Lead Financials)',             # 75
    'Average Deal Amount',                                     # 76
    'Conversion Rate (%)',                                     # 77
    'ROAS',                                                    # 78
    'CAC',                                                     # 79
    'A1',                                                      # 80
    'Source/Brand Breakdown',                                  # 81
]

_M, _I, _P = "money", "int", "pct"

# (published ledger label | None, sheet label, nth copy of that label on the
#  live tab, field, Q1, Q2, format)
_LEDGER = [
    # --- Budget & Efficiency
    ("Budget", "Budget", 1, "budget", 248446.0, 199796.0, _M),
    ("Spend", "Spend", 1, "spend", 239581.57, 187299.79, _M),
    ("Leads", "Leads", 1, "leads", 1281, 1220, _I),
    ("Qualified Leads", "Qualified Leads", 1, "qualified_leads", 636, 515, _I),
    ("Qualified Lead Ratio", "Qualified Lead Ratio", 1,
     "qualified_lead_ratio_pct", 49.65, 42.21, _P),
    ("Cost per Lead", "Cost per Lead", 1, "cost_per_lead", 187.03, 153.52, _M),
    ("Cost per Qualified Lead", "Cost per Qualified Lead", 1,
     "cost_per_qualified_lead", 376.70, 363.69, _M),
    ("Qualified Demos Booked (SDR+VAPI+Direct)",
     "Qualified Demos Booked (SDR+VAPI+Direct)", 1, "qual_demos_booked", 431, 343, _I),
    ("Total Demos Completed (SDR+VAPI+Direct)", "Demos Completed (SDR+VAPI+Direct)", 1,
     "demos_completed", 272, 238, _I),
    ("Show-up Rate (SDR+VAPI+Direct)", "Total Show Up Rate (%) (SDR+VAPI+Direct)", 1,
     "show_up_rate_pct", 63.11, 69.39, _P),
    ("Lost DNC (Bad Lead)", "Lost DNC (Bad Lead)", 1, "lost_dnc_bad_lead", 124, 149, _I),
    ("Cost / Qualified Demo Booked (SDR+VAPI+Direct)",
     "Cost Per Qualified Demo Booked (SDR+VAPI+Direct)", 1,
     "cost_per_qual_demo_booked", 555.87, 546.06, _M),
    # The ledger says "Qualified", the arithmetic says otherwise: 239,581.57 /
    # 880.81 = 272.0, the ALL-IN completed count, so this row is the sheet's
    # "Cost Per Demo Completed (SDR+VAPI+Direct)".
    ("Cost / Qualified Demo Completed (SDR+VAPI+Direct)",
     "Cost Per Demo Completed (SDR+VAPI+Direct)", 1,
     "cost_per_demo_completed", 880.81, 786.97, _M),
    # Not a ledger ROW - the denominator the ledger's Conversion Rate is computed
    # on, which the report has to show for its two cards to reconcile.
    (None, "Total Demos Completed (Direct)", 1, "demos_completed_direct", 248, 212, _I),
    # --- Projected - Actualized
    ("Number of Projected New Clients (Actualized)",
     "Number of Projected New Clients (Actualized)", 1,
     "projected_new_clients", 118, 133, _I),
    ("Total Projected Services Sold (Actualized)",
     "Total Projected Services Sold (Actualized)", 1,
     "projected_services_sold", 156, 224, _I),
    ("Projected Total Amount Sold ($) (Actualized)",
     "Projected Total Amount Sold ($) Actualized", 1,
     "projected_amount_sold", 479608.30, 649575.09, _M),
    ("Projected MRR w/o Setup Fees (Actualized)",
     "Projected MRR from New Sales w/o set up fees (Actualized)", 1,
     "projected_mrr_without_setup_fee", 422607.70, 594269.84, _M),
    # --- Revenue - Actualized
    ("Number of Revenue Clients (Actualized)", "Number of Revenue Clients (Actualized)", 1,
     "revenue_clients", 48, 57, _I),
    ("Total Services Sold (Actualized)", "Total Services Sold (Actualized)", 1,
     "services_sold", 81, 135, _I),
    ("Revenue Amount Sold (Actualized)", "Revenue Amount Sold (Actualized)", 1,
     "revenue_amount_sold", 262947.70, 401614.09, _M),
    ("Revenue Amount Sold w/o Setup Fee (Actualized)",
     "Revenue Amount Sold w/o Setup Fee (Actualized)", 1,
     "revenue_amount_sold_without_setup_fee", 231098.70, 372182.84, _M),
    # --- Projected - Not Actualized (rows 59-62; anchor row 59)
    ("Number of Projected New Clients (Not Actualized)",
     "Number of Projected New Clients (Not Actualized)", 1,
     "projected_new_clients_not_actualized", 114, 136, _I),
    ("Total Services Sold (Not Actualized)", "Total Services Sold (Not Actualized)", 1,
     "services_sold_not_actualized", 149, 227, _I),
    ("Revenue Amount Sold ($) (Not Actualized)", "Revenue Amount Sold ($) (Not Actualized)", 1,
     "revenue_amount_sold_not_actualized", 452607.70, 660703.09, _M),
    ("Revenue Amount Sold w/o Setup Fee (Not Actualized)",
     "Revenue Amount Sold w/o Setup Fee (Not Actualized)", 1,
     "revenue_amount_sold_without_setup_fee_not_actualized", 397087.70, 603397.84, _M),
    # --- Paying - Not Actualized (rows 63-66; anchor row 63)
    ("Number of Paying New Clients (Not Actualized)",
     "Number of Paying New Clients (Not Actualized)", 1,
     "paying_new_clients", 52, 58, _I),
    ("Total Services Sold (Not Actualized)", "Total Services Sold (Not Actualized)", 2,
     "paying_services_sold", 90, 137, _I),
    ("Revenue Amount Sold (Not Actualized)", "Revenue Amount Sold (Not Actualized)", 1,
     "paying_revenue_amount_sold", 290691.70, 409704.59, _M),
    ("Revenue Amount Sold w/o Setup Fee (Not Actualized)",
     "Revenue Amount Sold w/o Setup Fee (Not Actualized)", 2,
     "paying_revenue_amount_sold_without_setup_fee", 256943.70, 378958.84, _M),
    # --- Inbound Sales Pipeline
    ("Revenue Amount Sold (Inbound Sales Pipeline)",
     "Revenue Amount Sold (Inbound Sales Pipeline)", 1,
     "inbound_pipeline_revenue_amount_sold", 448996.54, 588791.09, _M),
    # --- Goal & Blended Financials
    ("% of Revenue Target Goal (Not Actualized)", "Percentage of Revenue Target Goal", 1,
     "revenue_target_pct", 94.29, 123.50, _P),
    ("Revenue Sold Goal Amount", "Revenue Sold Goal Amount", 1,
     "revenue_sold_goal", 480000.0, 535000.0, _M),
    ("Average Deal Amount", "Average Deal Amount", 1,
     "average_deal_amount", 4064.48, 4884.02, _M),
    ("Conversion Rate (%)", "Conversion Rate (%)", 1,
     "conversion_rate_pct", 19.35, 26.89, _P),
    # The sheet's bare "ROAS" row is the actualized one: 262,947.70 / 239,581.57
    # = 109.75%, exactly what the ledger publishes as "ROAS - Actualized (%)".
    ("ROAS \u2014 Actualized (%)", "ROAS", 1, "roas_pct", 109.75, 214.42, _P),
    ("CAC", "CAC", 1, "cac", 4991.28, 3285.96, _M),
]

# The one ledger row with no sheet row of its own, derived from the components
# of the same period: revenue amount sold (not actualized) / spend.
_LEDGER_DERIVED = ("ROAS \u2014 Not Actualized (%)", "roas_not_actualized_pct", 188.92, 352.75)

# Performance vs Investment decoys, so "Performance first" stays provable.
_INVESTMENT_DECOY = {"Budget": ("$260,000.00", "$210,000.00"),
                     "Spend": ("$248,110.00", "$195,000.00")}


def _cell(value, fmt):
    if fmt == _I:
        return f"{int(value):,}"
    if fmt == _P:
        return f"{value:.2f}%"
    return f"${value:,.2f}"


def _rollup_rows():
    """The live tab's rows, in the live order, carrying the ledger's figures.

    Every label the tab has is emitted, including the ones the ledger does not
    cover - blank, exactly as production has them. The blank stray copy at row
    74 is one of those, and it is the whole point: a fixture that drops the rows
    nobody reads cannot catch a read landing on one of them.
    """
    values = {}
    for _pub, label, nth, _field, q1, q2, fmt in _LEDGER:
        values[(label.strip().lower(), nth)] = (_cell(q1, fmt), _cell(q2, fmt))
    rows = [["All", "Jan (Performance)", "Jan (Investment)",
             "Feb (Performance)", "Feb (Investment)", "1st Quarter"]]
    seen = {}
    for label in _LIVE_LABELS:
        key = label.strip().lower()
        seen[key] = seen.get(key, 0) + 1
        perf1, perf2 = values.get((key, seen[key]), ("", ""))
        inv1, inv2 = _INVESTMENT_DECOY.get(label, ("", ""))
        rows.append([label, perf1, inv1, perf2, inv2, "x"])
    return rows


def _relabel(rows, old_label, new_label, occurrence=1):
    """Copy of *rows* with the *occurrence*-th ``old_label`` row renamed - how a
    row goes missing in real life (the team renames it), rather than deleted."""
    out, hits = [], 0
    for row in rows:
        if row and row[0] == old_label:
            hits += 1
            if hits == occurrence:
                row = [new_label] + row[1:]
        out.append(row)
    return out


def _with_cells(rows, label, perf, inv, occurrence=1):
    """Copy of *rows* with the *occurrence*-th *label* row given these Jan cells."""
    out, hits = [], 0
    for row in rows:
        if row and row[0] == label:
            hits += 1
            if hits == occurrence:
                row = [label, perf, inv] + row[3:]
        out.append(row)
    return out


def test_the_fixture_is_the_live_tab():
    """Guards the guard: if this fixture stops matching the dumped production
    label sequence, every assertion below is testing something that does not
    exist. The three duplicate/stray positions are named explicitly because they
    are what the parser has to survive."""
    assert len(_LIVE_LABELS) == 81
    assert _LIVE_LABELS[58] == "Number of Projected New Clients (Not Actualized)"  # row 59
    assert _LIVE_LABELS[59] == "Total Services Sold (Not Actualized)"              # row 60
    assert _LIVE_LABELS[60] == "Revenue Amount Sold ($) (Not Actualized)"          # row 61
    assert _LIVE_LABELS[62] == "Number of Paying New Clients (Not Actualized)"     # row 63
    assert _LIVE_LABELS[63] == "Total Services Sold (Not Actualized)"              # row 64
    assert _LIVE_LABELS[64] == "Revenue Amount Sold (Not Actualized)"              # row 65
    # The two revenue-amount rows are DIFFERENT strings - not a shared label.
    assert _LIVE_LABELS[60] != _LIVE_LABELS[64]
    # ...and the w/o-setup-fee label really does appear three times.
    assert _LIVE_LABELS.count("Revenue Amount Sold w/o Setup Fee (Not Actualized)") == 3
    assert _LIVE_LABELS[73] == "Revenue Amount Sold w/o Setup Fee (Not Actualized)"  # row 74


def test_both_revenue_amount_rows_resolve_from_their_own_block():
    """The production defect, pinned. One entry carrying both spellings as
    candidates resolved NEITHER row, because occurrence counts per candidate and
    never across the tuple - so occurrence 2 found nothing and the guard then
    correctly withheld occurrence 1 as well. Two live figures and the ROAS
    derived from one of them simply vanished from the payload."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    jan, feb = got["2026-01"], got["2026-02"]

    assert (jan["revenue_amount_sold_not_actualized"],
            feb["revenue_amount_sold_not_actualized"]) == (452607.70, 660703.09)
    assert (jan["paying_revenue_amount_sold"],
            feb["paying_revenue_amount_sold"]) == (290691.70, 409704.59)
    # And the field that depends on the first of them is back.
    assert (jan["roas_not_actualized_pct"], feb["roas_not_actualized_pct"]) == (188.92, 352.75)


def test_a_stray_copy_outside_the_blocks_is_never_read():
    """"...w/o Setup Fee (Not Actualized)" exists three times: Projected (62),
    Paying (66) and a blank stray at row 74. Lose the Projected copy and pure
    occurrence counting slides everything up - occurrence 1 serves the PAYING
    figure under the Projected name and occurrence 2 lands on the blank stray.
    Reading each row from its own block window instead, the Projected field goes
    absent and the Paying field is untouched."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    NO_FEE = "Revenue Amount Sold w/o Setup Fee (Not Actualized)"
    # Give the stray at row 74 a value it could never legitimately supply, so a
    # read landing there is loud rather than silently blank.
    rows = _with_cells(_rollup_rows(), NO_FEE, "$999,999.99", "", occurrence=3)
    # Then lose the Projected copy - the move that slides occurrence counting.
    rows = _relabel(rows, NO_FEE, "Revenue Amount Sold w/o Setup Fee (Signed)", occurrence=1)

    jan = parse_official_totals(rows, 2026)["2026-01"]
    assert "revenue_amount_sold_without_setup_fee_not_actualized" not in jan
    assert jan["paying_revenue_amount_sold_without_setup_fee"] == 256943.70
    assert 999999.99 not in jan.values()   # the stray is outside every window


def test_a_missing_block_header_makes_its_rows_absent_not_guessed():
    """The anchor IS the disambiguation. Rename the header that opens a section
    and its repeated rows can no longer be told from the other section's - so
    they go absent, and the other section keeps reading correctly."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    rows = _relabel(_rollup_rows(), "Number of Projected New Clients (Not Actualized)",
                    "Projected New Clients")
    jan = parse_official_totals(rows, 2026)["2026-01"]
    for field in ("projected_new_clients_not_actualized", "services_sold_not_actualized",
                  "revenue_amount_sold_not_actualized",
                  "revenue_amount_sold_without_setup_fee_not_actualized",
                  "roas_not_actualized_pct"):
        assert field not in jan, field
    # The Paying block is unaffected - its own anchor is still there.
    assert jan["paying_services_sold"] == 90.0
    assert jan["paying_revenue_amount_sold"] == 290691.70
    assert jan["paying_revenue_amount_sold_without_setup_fee"] == 256943.70


def test_losing_one_block_leaves_the_other_reading_correctly():
    """Rename the Projected copy of a genuinely shared label. The Projected field
    goes absent; the Paying field keeps its own figure instead of quietly
    inheriting the survivor - which is what first-match-wins did, and what pure
    occurrence counting does from the other direction."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    rows = _relabel(_rollup_rows(), "Total Services Sold (Not Actualized)",
                    "Total Services Sold (Projected)", occurrence=1)
    jan = parse_official_totals(rows, 2026)["2026-01"]
    assert "services_sold_not_actualized" not in jan
    assert jan["paying_services_sold"] == 90.0          # never the Projected 149

    rows = _relabel(_rollup_rows(), "Total Services Sold (Not Actualized)",
                    "Total Services Sold (Paying)", occurrence=2)
    jan = parse_official_totals(rows, 2026)["2026-01"]
    assert "paying_services_sold" not in jan
    assert jan["services_sold_not_actualized"] == 149.0  # never the Paying 90


def test_management_fees_is_absent_while_the_row_is_blank():
    """Label-correct at row 1 but both cells empty since July 2026 (last value
    June, $8,200.00). An unbilled month must read as absent - a $0 management
    fee is a claim about the contract, not a gap in the sheet."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    jan = parse_official_totals(_rollup_rows(), 2026)["2026-01"]
    assert "management_fees" not in jan
    assert 0 not in jan.values()

    # When it IS billed it resolves off the Investment column, Performance blank.
    billed = _with_cells(_rollup_rows(), "Management Fees", "", "$8,200.00")
    assert parse_official_totals(billed, 2026)["2026-01"]["management_fees"] == 8200.0


def test_every_ledger_row_has_a_parseable_source():
    """The ledger is the contract: every row the report renders must resolve to
    a field, from the right block, with the published figure. A row that goes
    unmapped is a card the report cannot fill - named here rather than found on
    a board slide."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    q1, q2 = got["2026-01"], got["2026-02"]

    unresolved = []
    for published, sheet_label, nth, field, want1, want2, _fmt in _LEDGER:
        if published is None:
            continue  # not a ledger row; asserted on its own below
        if q1.get(field) != want1 or q2.get(field) != want2:
            unresolved.append(
                f"{published!r} -> {field} (sheet row {sheet_label!r}, copy #{nth}): "
                f"got {q1.get(field)}/{q2.get(field)}, want {want1}/{want2}")
    assert not unresolved, "ledger rows with no parseable source:\n" + "\n".join(unresolved)

    published, field, want1, want2 = _LEDGER_DERIVED
    assert (q1[field], q2[field]) == (want1, want2), f"{published} is not derived correctly"

    # 37 ledger rows: 36 read from a sheet row, 1 derived.
    assert len([r for r in _LEDGER if r[0] is not None]) == 36


def test_conversion_rate_denominator_is_the_direct_demo_count():
    """The report showed Conversion Rate and Total Demos Completed as adjacent
    cards, and 48 / 272 = 17.65%, not the 19.35% printed beside it. The real
    denominator is the DIRECT completed count - 248, a number that appeared
    nowhere on the page. Parsing it is what lets the reader do the division."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    for key, demos in (("2026-01", 248), ("2026-02", 212)):
        m = got[key]
        assert m["demos_completed_direct"] == demos
        assert round(m["revenue_clients"] / m["demos_completed_direct"] * 100, 2) == \
            m["conversion_rate_pct"]
        assert m["demos_completed"] != m["demos_completed_direct"]
        assert round(m["revenue_clients"] / m["demos_completed"] * 100, 2) != \
            m["conversion_rate_pct"]


def test_cost_per_demo_completed_is_the_all_in_row_the_ledger_mislabels():
    """The ledger row reads "Cost / Qualified Demo Completed (SDR+VAPI+Direct)",
    but 239,581.57 / 880.81 = 272.0 - the ALL-IN completed count, not a qualified
    subset. So it is the sheet's plain cost-per-demo-completed row and needs no
    separate qualified field. Pinned because the label invites one."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    for key in ("2026-01", "2026-02"):
        m = got[key]
        assert round(m["spend"] / m["cost_per_demo_completed"]) == m["demos_completed"]


def test_the_revenue_goal_is_read_every_period_never_pinned():
    """The goal MOVES between periods (480,000 -> 535,000). A constant anywhere
    in the stack would silently misstate % of target for one of them."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    assert got["2026-01"]["revenue_sold_goal"] == 480000.0
    assert got["2026-02"]["revenue_sold_goal"] == 535000.0
    for key in ("2026-01", "2026-02"):
        m = got[key]
        assert round(m["revenue_amount_sold_not_actualized"] / m["revenue_sold_goal"] * 100, 2) == \
            m["revenue_target_pct"]


def test_official_totals_read_the_paying_block_not_the_projected_one():
    """Every "(Not Actualized)" figure must come from its own block. Crossing
    them duplicates the Projected numbers into the Paying rows - a wrong number
    that looks entirely plausible on the console."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    jan, feb = got["2026-01"], got["2026-02"]

    assert (jan["paying_services_sold"], feb["paying_services_sold"]) == (90.0, 137.0)
    assert (jan["paying_revenue_amount_sold"],
            feb["paying_revenue_amount_sold"]) == (290691.70, 409704.59)
    assert (jan["paying_revenue_amount_sold_without_setup_fee"],
            feb["paying_revenue_amount_sold_without_setup_fee"]) == (256943.70, 378958.84)

    assert (jan["services_sold_not_actualized"],
            feb["services_sold_not_actualized"]) == (149.0, 227.0)
    assert (jan["revenue_amount_sold_not_actualized"],
            feb["revenue_amount_sold_not_actualized"]) == (452607.70, 660703.09)
    assert (jan["revenue_amount_sold_without_setup_fee_not_actualized"],
            feb["revenue_amount_sold_without_setup_fee_not_actualized"]) == (397087.70, 603397.84)

    for paying, projected in (
        ("paying_services_sold", "services_sold_not_actualized"),
        ("paying_revenue_amount_sold", "revenue_amount_sold_not_actualized"),
        ("paying_revenue_amount_sold_without_setup_fee",
         "revenue_amount_sold_without_setup_fee_not_actualized"),
    ):
        assert jan[paying] != jan[projected] and feb[paying] != feb[projected]


def test_official_totals_extended_fields_keep_the_performance_first_basis():
    """The added rows read the same way the original eight do: Performance
    column first, Investment only as the fallback."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    jan = parse_official_totals(_rollup_rows(), 2026)["2026-01"]
    assert jan["budget"] == 248446.0              # Performance, not 260,000
    assert jan["spend"] == 239581.57              # Performance, not 248,110
    assert jan["lost_dnc_bad_lead"] == 124.0
    assert jan["inbound_pipeline_revenue_amount_sold"] == 448996.54
    assert jan["paying_new_clients"] == 52.0


def test_a_renamed_official_row_becomes_absent_never_zero():
    """Safe-by-construction: a removed or renamed sheet row must drop OUT of the
    payload. A zero would read as "the team sold nothing", which is a different
    and much more expensive claim than "the sheet no longer has this row"."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    rows = _relabel(_rollup_rows(), "Qualified Leads", "Qualified Leads (new)")
    jan = parse_official_totals(rows, 2026)["2026-01"]
    assert "qualified_leads" not in jan
    assert jan["leads"] == 1281.0           # its neighbours are unaffected
    assert all(v != 0 for v in jan.values())


def test_roas_not_actualized_is_derived_per_month_never_averaged():
    """No sheet row carries it: it is revenue_amount_sold_not_actualized / spend,
    recomputed from THAT period's own components. Averaging monthly ratios
    instead of recomputing them measures +1.74pp of error on this data."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    got = parse_official_totals(_rollup_rows(), 2026)
    jan, feb = got["2026-01"], got["2026-02"]

    assert jan["roas_not_actualized_pct"] == 188.92          # 452,607.70 / 239,581.57
    assert feb["roas_not_actualized_pct"] == 352.75          # 660,703.09 / 187,299.79
    for m in (jan, feb):
        assert round(m["revenue_amount_sold_not_actualized"] / m["spend"] * 100, 2) == \
            m["roas_not_actualized_pct"]
    mean = round((jan["roas_not_actualized_pct"] + feb["roas_not_actualized_pct"]) / 2, 2)
    assert jan["roas_not_actualized_pct"] != feb["roas_not_actualized_pct"] != mean
    assert (jan["roas_pct"], feb["roas_pct"]) == (109.75, 214.42)


def test_roas_not_actualized_is_absent_when_an_input_is():
    """Same missing-field law as the parsed rows - a derived field with a missing
    input is absent, not 0. A 0% ROAS on a board report is a fire drill."""
    from marketing_research_agent.sources.sheets_source import parse_official_totals

    no_revenue = _relabel(_rollup_rows(), "Revenue Amount Sold ($) (Not Actualized)",
                          "Revenue Amount Sold ($) (Signed)")
    jan = parse_official_totals(no_revenue, 2026)["2026-01"]
    assert "revenue_amount_sold_not_actualized" not in jan
    assert "roas_not_actualized_pct" not in jan
    # The Paying row of the same family is untouched by its neighbour's rename.
    assert jan["paying_revenue_amount_sold"] == 290691.70

    no_spend = _relabel(_rollup_rows(), "Spend", "Spend (media)")
    jan = parse_official_totals(no_spend, 2026)["2026-01"]
    assert "spend" not in jan and "roas_not_actualized_pct" not in jan

    zero_spend = _with_cells(_rollup_rows(), "Spend", "$0.00", "")
    jan = parse_official_totals(zero_spend, 2026)["2026-01"]
    assert jan["spend"] == 0.0 and "roas_not_actualized_pct" not in jan


def test_row_for_occurrence_defaults_to_the_first_match():
    """The default keeps every pre-existing caller (parse_tracker, the vendor
    field map) resolving exactly the row it always resolved."""
    from marketing_research_agent.sources.sheets_source import _row_for

    rows = [["All"], ["Spend", "1"], ["Leads", "2"], ["Spend", "3"]]
    assert _row_for(rows, 1, len(rows), ["spend"]) == 1          # default = 1st
    assert _row_for(rows, 1, len(rows), ["spend"], 1) == 1
    assert _row_for(rows, 1, len(rows), ["spend"], 2) == 3
    assert _row_for(rows, 1, len(rows), ["spend"], 3) is None    # absent, not the last one
    assert _row_for(rows, 1, len(rows), ["total spend", "spend"], 2) == 3


def test_the_original_eight_official_fields_are_untouched():
    """The extension is additive. These eight keys, their label lists and their
    order are what every current caller (headline strip, trends, reconciliation)
    reads, so they are pinned verbatim."""
    from marketing_research_agent.sources.sheets_source import _OFFICIAL_FIELD_LABELS

    legacy = {
        "budget": ["budget"],
        "spend": ["spend"],
        "leads": ["leads"],
        "qualified_leads": ["qualified leads"],
        "demos_booked": ["total demos booked (sdr+vapi+direct)", "total demos booked"],
        "qual_demos_booked": ["qualified demos booked (sdr+vapi+direct)",
                              "qualified demos booked"],
        "demos_completed": ["demos completed (sdr+vapi+direct)", "total demos completed"],
        "services_sold": ["total services sold (actualized)"],
    }
    for field, labels in legacy.items():
        got_labels, where = _OFFICIAL_FIELD_LABELS[field]
        assert list(got_labels) == labels
        assert where == 1
